# coding=utf-8

import datetime
import time

# import xlwt
# from xlwt import *
from openpyxl import Workbook
from openpyxl.styles import NamedStyle, Font, Border, Side, PatternFill
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font

from statistics.config import AreaTool, ProvinceTool, ADS_FIELD, CERT_CONSTANT, TITLE_DICT
from statistics.dao import SsdbOrderDao, AjhzOrderDao, LawyerCaseDao, LawyerDao

YEAR = datetime.datetime.now().year  # 当前年
MONTH = datetime.datetime.now().month  # 当前月
DAY = datetime.datetime.now().day  # 当前日
# workbook = xlwt.Workbook(encoding='utf-8', style_compression=0)
# sheet = workbook.add_sheet('认证律师用户画像', cell_overwrite_ok=True)
# style = XFStyle()  # 赋值style为XFStyle()，初始化样式
# font = xlwt.Font()  # 字体基本设置
# font.name = u'Cambria'
# font.color = 'black'
# font.height = 220  # 字体大小，220就是11号字体，大概就是11*20得来的吧
# style.font = font
#
# borders = xlwt.Borders()
# borders.left = 1
# borders.right = 1
# borders.top = 1
# borders.bottom = 1
# borders.bottom_colour = 0x3A
# style.borders = borders

# ------------------


# ----------------
from openpyxl.styles import Border, Side, Font  # 设置字体和边框需要的模块

# 设置字体样式
font = Font(u'Cambria', size=11, bold=False, italic=False, strike=False, color='000000')
# 设置边框样式，上下左右边框
border = Border(left=Side(style='thin', color='000000'), right=Side(style='thin', color='000000'),
                top=Side(style='thin', color='000000'), bottom=Side(style='thin', color='000000'))


# --------------

class CaseBean(object):
    def __init__(self, kwargs):
        self.total_case_num = 0
        self.ads_field = {}
        for key, value in ADS_FIELD.items():
            match = False
            count = 0
            for data in kwargs:
                if data:
                    caseType = data["caseType"]
                    if key in caseType:
                        count = data["count"]
                        break
            self.ads_field[value] = count
            self.total_case_num += count

    def __str__(self):
        return str(self.__dict__)

    def __repr__(self):
        return self.__str__()


class LawyerBean(object):
    def __init__(self, *args, **kwargs):
        self.lawyer_id = kwargs.get("id", "")
        self.lawyer_name = kwargs.get("real_name", "")
        self.phone = kwargs.get("phone", "")
        self.device_type = kwargs.get("deviceType", "")
        self.device_info = kwargs.get("deviceInfo", "")
        self.deposit_status = kwargs.get("depositStatus", "")
        self.lawyer_type = kwargs.get("lawyerType", "")
        self.proceed_lawyer_type()
        self.create_date = kwargs.get("create_date", "")
        self.regist_year = ""
        if (isinstance(self.create_date, datetime.datetime)):
            self.regist_year = self.create_date.year  # 注册年份
        self.proceed_deposit_status()
        self.id_card_no = kwargs.get("idCard", "")
        self.proceed_id_card_no()
        self.cert_num = kwargs.get("certNum", "")
        self.proceed_cert_num()
        self.province_id = kwargs.get("province_id", "")
        self.city_id = kwargs.get("city_id", "")
        self.status = kwargs.get("status", 0)
        self.proceed_status()
        self.proceed_work_space()
        self.case_bean = CaseBean([])
        self.proceed_case()
        self.ssdb_order_num = 0
        self.proceed_ssdb_order()
        self.ajhz_order_distribute_num = 0  # 发单
        self.ajhz_order_accept_num = 0  # 接单
        self.proceed_ajhz_order()

    def is_second_generation_identity_card(self):
        return self.id_card_no and len(self.id_card_no) == 18

    def proceed_lawyer_type(self):
        lawyer_type = self.lawyer_type
        lawyer_type_desc = ""
        if lawyer_type == 0:
            lawyer_type_desc = "执业律师"
        elif lawyer_type == 1:
            lawyer_type_desc = "实习律师"
        self.lawyer_type_desc = lawyer_type_desc

    def is_common_cert_num(self):
        return self.cert_num and len(self.cert_num) == 17

    def proceed_status(self):
        _status = str(self.status)
        # 0：注册
        # 1：已经完善资料
        # 2：认证审核中
        # 3：认证审核失败
        # 4: 认证成功（已认证）
        _ret = "注册"
        if "0" in _status:
            _ret = "注册"
        elif "1" in _status:
            _ret = "已经完善资料"
        elif "2" in _status:
            _ret = "认证审核中"
        elif "3" in _status:
            _ret = "认证失败"
        elif "4" in _status:
            _ret = "认证成功"
        self.status_desc = _ret

    def proceed_ssdb_order(self):
        row = SsdbOrderDao.count(self.lawyer_id)
        self.ssdb_order_num = row.get("count", 0)

    def proceed_ajhz_order(self):
        self.ajhz_order_distribute_num = AjhzOrderDao.distribute_count(self.lawyer_id).get("count", 0)
        self.ajhz_order_accept_num = AjhzOrderDao.accept_count(self.lawyer_id).get("count", 0)

    def proceed_case(self):
        row = LawyerCaseDao.group(self.lawyer_id)
        self.case_bean = CaseBean(row)

    def proceed_deposit_status(self):
        deposit_desc = ""
        _deposit_status = str(self.deposit_status)
        if "-1" in _deposit_status:
            deposit_desc = "未付保证金"
        elif "0" in _deposit_status:
            deposit_desc = "已付保证金"
        elif "1" in _deposit_status:
            deposit_desc = "申请退保证金"
        elif "2" in _deposit_status:
            deposit_desc = "已退保证金"
        self.deposit_desc = deposit_desc

    def proceed_cert_num(self):
        self.cert_year = ""
        self.cert_now_year = ""
        self.cert_type = ""

        if self.is_common_cert_num():
            __cert_type = self.cert_num[9:10]  # 执业类别
            __cert_year = self.cert_num[5:9]
            self.cert_year = __cert_year
            self.cert_now_year = YEAR - int(__cert_year) + 1
            self.cert_type = CERT_CONSTANT.get(__cert_type)

    def proceed_work_space(self):
        self.province_name = 0
        self.city_name = 0
        if self.province_id:
            self.province_name = ProvinceTool.PROVINCE_DICT.get(str(self.province_id), self.province_id)
        if self.city_id:
            self.city_name = ProvinceTool.CITY_DICT.get(str(self.city_id), self.city_id)

    def proceed_id_card_no(self):
        self.provice = ""
        self.city = ""
        self.town = ""
        self.sex = ""
        self.age = ""
        if self.is_second_generation_identity_card():
            print(self.id_card_no)
            birth_year = self.id_card_no[6:10]
            self.age = YEAR - int(birth_year)
            sex_num = self.id_card_no[16:17]  # 第17位数字表示性别：奇数表示男性，偶数表示女性
            if (int(sex_num) % 2) == 0:
                self.sex = "女"
            else:
                self.sex = "男"
            provice = self.id_card_no[0:2]
            city = self.id_card_no[2:4]
            town = self.id_card_no[4:6]
            self.provice = AreaTool.query_name(provice)
            self.city = AreaTool.query_name(provice, city)
            self.town = AreaTool.query_name(provice, city, town)

    def __str__(self):
        return str(self.__dict__)

    def __repr__(self):
        return self.__str__()


class TitleRow(object):
    ROW = 1  # 标题2行

    def __init__(self, sheet, **kwargs):
        self.sheet = sheet
        index = 1
        sheet.font = font
        sheet.border = border

        for key, value in kwargs.items():
            start_index = index
            if isinstance(value, list) and value:
                for sub_value in value:
                    _cell = sheet.cell(row=2, column=index, value=sub_value)
                    _cell.border = border
                    _cell.font = font
                    index += 1
                print(start_index, index, key, "--------")
                sheet.merge_cells(start_row=1, start_column=start_index, end_row=1, end_column=index - 1)
                _cell = sheet.cell(row=1, column=start_index, value=key)
                _cell.border = border
                _cell.font = font
            else:
                print(start_index, index, key, "========")
                sheet.merge_cells(start_row=1, start_column=index, end_row=2, end_column=index)
                _cell = sheet.cell(row=1, column=index, value=key)
                _cell.border = border
                _cell.font = font
                index += 1

    def write(self, bean: LawyerBean):
        self.ROW += 1
        if bean:
            data = []
            data.append(bean.lawyer_id)
            data.append(bean.lawyer_name)
            data.append(bean.phone)
            data.append(bean.age)
            data.append(bean.sex)
            data.append(bean.id_card_no)
            data.append(bean.lawyer_type_desc)
            data.append(bean.regist_year)
            data.append(bean.status_desc)
            data.append(bean.cert_now_year)
            data.append(bean.cert_type)
            data.append(bean.cert_num)
            data.append(bean.province_name)
            data.append(bean.city_name)
            data.append("")
            data.append(bean.provice)
            data.append(bean.city)
            data.append(bean.town)
            data.append(bean.ssdb_order_num)
            data.append(bean.ajhz_order_distribute_num)
            data.append(bean.ajhz_order_accept_num)
            data.append(bean.device_type)
            data.append(bean.device_info)
            data.append(bean.case_bean.total_case_num)
            for key, value in bean.case_bean.ads_field.items():
                data.append(value)
            self.sheet.append(data)
            row = self.sheet.row_dimensions[1]
            row.font = font
            row.border = border


# lawyer = LawyerBean(**{'id': '4a423e0c02b9e7aa5acf2b4d06427ac1', 'address': '广州市天河路101号兴业银行大厦十三楼',
#                        'briefInfo': '国信信扬律师事务所的合伙人。\n专职律师，擅长劳动法、公司法等领域的法律事务。\n同时担任广州市律师协会劳动和社会保障法律专业委员会委员、广东省律师协会劳动和社会保障法律专业委员会委员。先后担任广东省飞来峡水利枢纽管理局、广州市国土资源和房屋管理局越秀分局、广州市越秀区旧城改造项目办公室、广州市越秀区房屋租赁管理所、马士基信息处理(广东)有限公司、卓域集团有限公司、金源矿业有限公司等多个政府单位及大型企业的常年法律顾问。',
#                        'caseNum': 101, 'license': '/certificate/2016/1/3/13c190a132414708b4f76547ad33917e.jpg',
#                        'certificateDate': datetime.datetime(2006, 6, 1, 0, 0), 'city_id': 1292,
#                        'create_date': datetime.datetime(2014, 12, 15, 0, 17, 12), 'deviceType': 'android',
#                        'email': '17960413@qq.com', 'login_name': '13076828028',
#                        'logo': '/logo/lawyer/2016/1/12/4a423e0c02b9e7aa5acf2b4d06427ac1.jpg', 'phone': '13076828028',
#                        'moneyBlocked': 0.0, 'moneyOut': 0.0, 'moneySum': 30.049999999999997, 'law_firm': '国信信扬律师事务所',
#                        'orderFlag': 1, 'password': 'fa33ab998750389624466dd272cbecaf', 'province_id': 29,
#                        'real_name': '冯志成', 'recommend': 0, 'sex': 0, 'status': 4,
#                        'update_date': datetime.datetime(2018, 12, 4, 14, 48, 36), 'weixin': '13076828028',
#                        'evaluate_id': '4a423e0c02b9e7aa5acf2b4d06427ac1',
#                        'applyDate': datetime.datetime(2016, 1, 14, 10, 25, 51), 'failType': None,
#                        'certNum': '14401200610938444', 'points': 0, 'pointsExchange': 0, 'lawyerType': 0,
#                        'appVersion': '5.9.1', 'deviceId': '4C643EDD-36EC-487B-80FD-E4E79325F69A',
#                        'deviceInfo': 'Simulator,9.30', 'idCard': '410225198912284693',
#                        'dwlsd_openid': 'oLn3Vs6NC9VMlM-Q1TvTqsRUngMo', 'moneyTj': None, 'moneyTz': None,
#                        'deposit': None,
#                        'depositStatus': -1, 'pointsBlocked': 0, 'pointsOut': 1000, 'pointsSum': 100000, 'channel': None,
#                        'depositBackDate': None, 'stateJoin': 0, 'security': 0, 'advJoinState': 1, 'wxPush': 1,
#                        'wxPushOpenId': 'oLn3Vszm0ylUD2RQqt_2fGitTq50', 'wxSsdbOpenId': None})

##################excel 标题配置 开始#####################

# lawyer_id_list = LawyerDao.extract_all_lawyer_id()
lawyer_id_list = [{'id': '4a423e0c02b9e7aa5acf2b4d06427ac1'}]
_data = LawyerDao.query_by_lawyer_id(lawyer_id_list[0].get("id"))
lawyer = LawyerBean(**_data)
# ------------------
wb = Workbook()
ws = wb.active
ws.title = "认证律师用户画像"
print(lawyer_id_list)
row = TitleRow(sheet=ws, **TITLE_DICT)
print(lawyer)
row.write(lawyer)
wb.save('.\{}统计表_data.xlsx'.format(str(time.time())))
# index = 0
# for lawyer in lawyer_id_list:
#     _data = LawyerDao.query_by_lawyer_id(lawyer.get("id"))
#     lawyer = LawyerBean(**_data)
#     row.write(lawyer)
#     index += 1
#     if index % 10000 == 0:
#         print("==========save start===================")
#         workbook.save(r'.\{}统计表.xlxs'.format(str(time.time())))
#         print("==========save ok===================")
#     print("休眠1秒钟 ", str(index))
# workbook.save(r'.\{}统计表_end.xls'.format(str(time.time())))
##################excel 标题配置 结束#####################
# ---------------------------
# sheet = workbook.add_sheet('律师案例数据分析', cell_overwrite_ok=True)
# title_dict = {
#     "律师ID": "",
#     "律师姓名": "",
#     "律师手机号": "",
#     "年龄": "11",
#     "性别": "22",
#     "执业地区": "",
#     "家乡地区": "",
#     "执业年限": "",
#     "案例数量": "",
#     "诉保订单": "",
#     "案件合作": "",
#     "保证金": "",
#     "用户机型": "",
#     "执业类别": "",
# }
# # ---------------------------
# workbook.save(r'.\统计表.xls')
